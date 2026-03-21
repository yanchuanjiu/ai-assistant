"""Excel 文件下载与解析工具。支持飞书 IM 消息文件、飞书云盘文件，处理合并单元格。"""
import io
import json
import hashlib
import os
import time
import logging

logger = logging.getLogger(__name__)

TEMP_DIR = "/tmp/excel_import"
os.makedirs(TEMP_DIR, exist_ok=True)

# session 有效期 2 小时
SESSION_TTL = 7200


def _session_path(session_key: str) -> str:
    return os.path.join(TEMP_DIR, f"{session_key}.json")


def save_session(data: dict) -> str:
    """将解析结果存入临时文件，返回 session_key。"""
    raw = f"{time.time()}-{os.getpid()}"
    key = hashlib.md5(raw.encode()).hexdigest()[:10]
    payload = {"ts": time.time(), "data": data}
    with open(_session_path(key), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    return key


def load_session(session_key: str) -> dict | None:
    """读取临时文件中的解析结果，不存在或已过期返回 None。"""
    path = _session_path(session_key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        if time.time() - payload.get("ts", 0) > SESSION_TTL:
            os.remove(path)
            return None
        return payload["data"]
    except Exception:
        return None


# --------------------------------------------------------------------------- #
# 下载
# --------------------------------------------------------------------------- #

def download_feishu_im_file(message_id: str, file_key: str) -> bytes:
    """从飞书 IM 消息中下载文件。"""
    import httpx
    from integrations.feishu.client import get_tenant_access_token, FEISHU_BASE
    token = get_tenant_access_token()
    resp = httpx.get(
        f"{FEISHU_BASE}/im/v1/messages/{message_id}/resources/{file_key}",
        headers={"Authorization": f"Bearer {token}"},
        params={"type": "file"},
        timeout=60,
        follow_redirects=True,
    )
    resp.raise_for_status()
    return resp.content


def download_feishu_drive_file(file_token: str) -> bytes:
    """从飞书云盘下载文件（优先 user token，降级 tenant token）。"""
    import httpx
    from integrations.feishu.client import get_tenant_access_token, FEISHU_BASE
    try:
        from integrations.feishu.client import get_user_access_token
        token = get_user_access_token()
    except Exception:
        token = get_tenant_access_token()

    resp = httpx.get(
        f"{FEISHU_BASE}/drive/v1/files/{file_token}/download",
        headers={"Authorization": f"Bearer {token}"},
        timeout=60,
        follow_redirects=True,
    )
    resp.raise_for_status()
    return resp.content


# --------------------------------------------------------------------------- #
# 解析
# --------------------------------------------------------------------------- #

def _cell_value(cell):
    """提取单元格值，转为 JSON 可序列化类型。"""
    v = cell.value
    if v is None:
        return None
    if hasattr(v, "isoformat"):       # datetime / date / time
        return v.isoformat()
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return int(v)
        return round(v, 8)
    return v


def parse_excel_bytes(file_bytes: bytes, file_name: str = "file.xlsx") -> dict:
    """
    解析 Excel 字节流，处理合并单元格。

    返回：
    {
        "file_name": str,
        "sheets": [
            {
                "name": str,
                "rows": [[cell_value, ...], ...],   # 二维列表，合并单元格已展开
                "row_count": int,
                "col_count": int,
                "merged_cells": [str, ...]           # 合并信息摘要
            }
        ]
    }
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("缺少 openpyxl，请运行：pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result: dict = {"file_name": file_name, "sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 构建合并单元格映射：(row, col) -> value
        merged_map: dict = {}
        merged_info: list[str] = []
        for mc in ws.merged_cells.ranges:
            top_left = ws.cell(mc.min_row, mc.min_col)
            val = _cell_value(top_left)
            merged_info.append(f"{mc}='{val}'")
            for r in range(mc.min_row, mc.max_row + 1):
                for c in range(mc.min_col, mc.max_col + 1):
                    merged_map[(r, c)] = val

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        rows: list[list] = []
        for r in range(1, max_row + 1):
            row: list = []
            for c in range(1, max_col + 1):
                if (r, c) in merged_map:
                    row.append(merged_map[(r, c)])
                else:
                    row.append(_cell_value(ws.cell(r, c)))
            # 去除行尾空值
            while row and (row[-1] is None or row[-1] == ""):
                row.pop()
            rows.append(row)

        # 去除末尾空行
        while rows and all(v is None or v == "" for v in rows[-1]):
            rows.pop()

        result["sheets"].append({
            "name": sheet_name,
            "rows": rows,
            "row_count": len(rows),
            "col_count": max_col,
            "merged_cells": merged_info[:30],
        })

    return result


# --------------------------------------------------------------------------- #
# 预览
# --------------------------------------------------------------------------- #

def preview_excel(parsed: dict, max_rows: int = 5) -> str:
    """生成 Excel 解析结果的可读预览。"""
    lines = [f"📊 文件：{parsed['file_name']}，共 {len(parsed['sheets'])} 个工作表"]
    for idx, sheet in enumerate(parsed["sheets"]):
        lines.append(
            f"\n**[{idx}] 工作表：{sheet['name']}**"
            f"（{sheet['row_count']} 行 × {sheet['col_count']} 列）"
        )
        if sheet["merged_cells"]:
            mc_preview = ", ".join(sheet["merged_cells"][:5])
            lines.append(f"合并单元格（{len(sheet['merged_cells'])} 个）: {mc_preview}")

        rows = sheet["rows"]
        if not rows:
            lines.append("（空工作表）")
            continue

        header = rows[0]
        lines.append(f"第1行（表头）: {header}")

        preview_count = min(max_rows, len(rows) - 1)
        if preview_count > 0:
            lines.append(f"前 {preview_count} 行数据：")
            for i, row in enumerate(rows[1: 1 + preview_count], 1):
                lines.append(f"  行{i + 1}: {row}")

        if len(rows) > max_rows + 1:
            lines.append(f"  ...（共 {len(rows)} 行，含表头）")

    return "\n".join(lines)
